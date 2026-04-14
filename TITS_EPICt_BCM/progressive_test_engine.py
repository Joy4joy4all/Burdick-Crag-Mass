# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
PROGRESSIVE TEST ENGINE — The Missing Connector
======================================================
Ties test_intelligence + post_test_generator + Excel plan
into one progressive KNN-Bayesian flow.

FORWARD:  Click person → calculate(all_prior) → compound questions → template
BACKWARD: Doc reader done → calculate(all+new) → delta → slide → report → IoC
REVERSE:  Foreman ocular → what was missed → nothing to suggest

Each test_run save creates a dated IoC (Test Run of Change) snapshot.
KNN model builds progressively with every test_run.
Excel plan becomes a living progressive document.

© 2025-2026 Stephen J. Burdick Sr. — All Rights Reserved.
For all the industrial workers lost to preventable acts.
"""

import json, copy, hashlib
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

# ── Core engines ──
try:
    from test_intelligence import TestIntelligence
    INTEL_AVAILABLE = True
except ImportError:
    INTEL_AVAILABLE = False

try:
    from post_test_generator import PostTest RunGenerator, load_hypotheses_from_excel
    POST_GEN_AVAILABLE = True
except ImportError:
    POST_GEN_AVAILABLE = False

try:
    from project_paths import (
        get_project_dir, get_report_slide_dir, get_tested_dir,
        find_project_config, get_active_project, BCM_ROOT,
    )
    PATHS_AVAILABLE = True
except ImportError:
    PATHS_AVAILABLE = False
    def get_project_dir(pid): return Path.cwd()
    def get_active_project(): return "BCM_SUBSTRATE"
    BCM_ROOT = Path.cwd()

try:
    from sklearn.neighbors import NearestNeighbors
    from sklearn.feature_extraction.text import TfidfVectorizer
    import numpy as np
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False


# ══════════════════════════════════════════════════════════════
# IoC — TEST OF CHANGE SNAPSHOT
# ══════════════════════════════════════════════════════════════
class Test RunOfChange:
    """
    Dated snapshot of intelligence state after each test.
    Chain of IoC snapshots = progressive learning record.
    KNN operates over this chain to find nearest prior states.
    """

    @staticmethod
    def save(project_id: str, person: str,
             state_before: dict, state_after: dict, delta: dict,
             test_num: int = 0):
        """Save IoC snapshot to project's ioc/ folder."""
        ioc_dir = get_project_dir(project_id) / "ioc"
        ioc_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"IoC_{timestamp}_{person.replace(' ', '_')}.json"

        snapshot = {
            "timestamp": datetime.now().isoformat(),
            "test_num": test_num,
            "script_name": person,
            "project_id": project_id,
            "state_before": _serialize_state(state_before),
            "state_after": _serialize_state(state_after),
            "delta": delta,
            "checksum": hashlib.md5(
                json.dumps(_serialize_state(state_after), sort_keys=True).encode()
            ).hexdigest()[:12],
        }

        path = ioc_dir / filename
        path.write_text(json.dumps(snapshot, indent=2, default=str), encoding='utf-8')
        print(f"  [IoC] Saved: {filename}")
        return path

    @staticmethod
    def load_chain(project_id: str) -> List[dict]:
        """Load all IoC snapshots in chronological order."""
        ioc_dir = get_project_dir(project_id) / "ioc"
        if not ioc_dir.exists():
            return []
        files = sorted(ioc_dir.glob("IoC_*.json"))
        chain = []
        for f in files:
            try:
                chain.append(json.loads(f.read_text(encoding='utf-8')))
            except Exception as e:
                print(f"  [IoC WARN] Failed to load {f.name}: {e}")
        return chain

    @staticmethod
    def get_state_at(project_id: str, test_num: int = None,
                     person_name: str = None) -> Optional[dict]:
        """Get intelligence state as-of a specific test_run.

        Matches by person_name first, falls back to test_num for legacy.
        """
        chain = Test RunOfChange.load_chain(project_id)
        if person_name:
            pn = person_name.strip().lower()
            for snapshot in reversed(chain):
                if snapshot.get('script_name', '').strip().lower() == pn:
                    return snapshot.get('state_after', {})
        if test_num is not None:
            for snapshot in reversed(chain):
                if snapshot.get('test_num', 0) <= test_num:
                    return snapshot.get('state_after', {})
        return None


def _serialize_state(state: dict) -> dict:
    """Make state JSON-serializable."""
    if not state:
        return {}
    result = {}
    for k, v in state.items():
        if isinstance(v, (str, int, float, bool, type(None))):
            result[k] = v
        elif isinstance(v, (list, tuple)):
            result[k] = [_serialize_item(i) for i in v]
        elif isinstance(v, dict):
            result[k] = {str(kk): _serialize_item(vv) for kk, vv in v.items()}
        else:
            result[k] = str(v)
    return result


def _serialize_item(item):
    if isinstance(item, (str, int, float, bool, type(None))):
        return item
    elif isinstance(item, dict):
        return {str(k): _serialize_item(v) for k, v in item.items()}
    elif isinstance(item, (list, tuple)):
        return [_serialize_item(i) for i in item]
    return str(item)


# ══════════════════════════════════════════════════════════════
# DELTA CALCULATOR — Before/After comparison
# ══════════════════════════════════════════════════════════════
def compute_delta(state_before: dict, state_after: dict) -> dict:
    """
    Compute what changed between two intelligence states.
    This is the 'reverse ocular' — shows exactly what this test_run revealed.
    """
    if not state_before or not state_after:
        return {"no_prior_state": True}

    delta = {
        "test_count_change": (
            state_after.get('test_count', 0) -
            state_before.get('test_count', 0)
        ),
    }

    # VP confidence changes
    vp_before = state_before.get('vp_scores', {})
    vp_after = state_after.get('vp_scores', {})
    vp_changes = {}
    for vp_id in set(list(vp_before.keys()) + list(vp_after.keys())):
        b_conf = vp_before.get(vp_id, {}).get('confidence', 0)
        a_conf = vp_after.get(vp_id, {}).get('confidence', 0)
        b_status = vp_before.get(vp_id, {}).get('status', 'NO DATA')
        a_status = vp_after.get(vp_id, {}).get('status', 'NO DATA')
        if b_conf != a_conf or b_status != a_status:
            vp_changes[vp_id] = {
                'confidence_before': b_conf, 'confidence_after': a_conf,
                'confidence_delta': a_conf - b_conf,
                'status_before': b_status, 'status_after': a_status,
                'status_changed': b_status != a_status,
            }
    delta['vp_changes'] = vp_changes

    # Hypothesis changes
    hyp_before = state_before.get('hypotheses', {})
    hyp_after = state_after.get('hypotheses', {})
    hyp_changes = {}
    for h_id in set(list(hyp_before.keys()) + list(hyp_after.keys())):
        b_conf = hyp_before.get(h_id, {}).get('confidence', 0)
        a_conf = hyp_after.get(h_id, {}).get('confidence', 0)
        if b_conf != a_conf:
            hyp_changes[h_id] = {
                'confidence_before': b_conf, 'confidence_after': a_conf,
                'delta': a_conf - b_conf,
            }
    delta['hypothesis_changes'] = hyp_changes

    # Equipment ranking changes
    eq_before = {name: stats for name, stats in state_before.get('equipment_ranked', [])}
    eq_after = {name: stats for name, stats in state_after.get('equipment_ranked', [])}
    new_equipment = [name for name in eq_after if name not in eq_before]
    delta['new_equipment'] = new_equipment

    # Damage total change
    delta['total_damage_before'] = state_before.get('total_damage_documented', 0)
    delta['total_damage_after'] = state_after.get('total_damage_documented', 0)
    delta['damage_delta'] = delta['total_damage_after'] - delta['total_damage_before']

    # Cube density change
    delta['layer_density_before'] = state_before.get('layer_density', {})
    delta['layer_density_after'] = state_after.get('layer_density', {})

    # New contradictions
    c_before = set(c.get('probe', '') for c in state_before.get('contradictions', []))
    c_after = state_after.get('contradictions', [])
    delta['new_contradictions'] = [c for c in c_after if c.get('probe', '') not in c_before]

    # Cube gaps closed
    gaps_before = set(state_before.get('cube_gaps', []))
    gaps_after = set(state_after.get('cube_gaps', []))
    delta['gaps_closed'] = list(gaps_before - gaps_after)
    delta['gaps_remaining'] = list(gaps_after)

    return delta


# ══════════════════════════════════════════════════════════════
# KNN PROGRESSIVE MODEL
# ══════════════════════════════════════════════════════════════
class ProgressiveKNN:
    """
    KNN model over IoC chain. Finds tests most similar to a target
    position to inform progressive question generation.

    Feature vector per test:
      - Q-Cube position (one-hot: L1/L2/L3, OA/OB/OC)
      - VP confidence scores at time of test_run (6 dims)
      - Equipment mention count (1 dim)
      - Dollar capture (1 dim, normalized)
      - Completeness score (1 dim)
      = 17 dimensions
    """

    LAYERS = ['L1', 'L2', 'L3']
    OBJECTS = ['OA', 'OB', 'OC']
    VP_IDS = ['VP1_DAMAGE_PREVENTION', 'VP2_PLANNED_MAINTENANCE',
              'VP3_OPERATOR_VISIBILITY', 'VP4_CHEMICAL_OPTIMIZATION',
              'VP5_SUPPLIER_ACCOUNTABILITY', 'VP6_PITCH_CONTAMINATION']

    def __init__(self):
        self.vectors = []
        self.test_data = []
        self.model = None

    def fit_from_chain(self, ioc_chain: List[dict], test_runs: List[dict]):
        """Build KNN from IoC chain + test data."""
        if not SKLEARN_AVAILABLE:
            return

        self.vectors = []
        self.test_data = []

        for iv in test_runs:
            if not iv.get('results') or '[PENDING' in str(iv.get('results', '')):
                continue

            vec = self._vectorize(iv)
            self.vectors.append(vec)
            self.test_data.append(iv)

        if len(self.vectors) >= 2:
            X = np.array(self.vectors)
            k = min(3, len(X) - 1)
            self.model = NearestNeighbors(n_neighbors=k, metric='cosine')
            self.model.fit(X)

    def find_similar(self, target: dict, k: int = 3) -> List[Tuple[dict, float]]:
        """Find k most similar prior tests to target."""
        if not self.model or not SKLEARN_AVAILABLE:
            return []

        vec = self._vectorize(target)
        distances, indices = self.model.kneighbors([vec])

        results = []
        for dist, idx in zip(distances[0], indices[0]):
            results.append((self.test_data[idx], float(1 - dist)))
        return results

    def _vectorize(self, iv: dict) -> list:
        """Convert test_run to feature vector."""
        vec = []
        # Layer one-hot (3)
        layer = iv.get('q_layer', '')
        vec.extend([1.0 if layer == l else 0.0 for l in self.LAYERS])
        # Object one-hot (3)
        obj = iv.get('q_object', '')
        vec.extend([1.0 if obj == o else 0.0 for o in self.OBJECTS])
        # VP confidence at test_run time (6) — use keyword scan as proxy
        text = ' '.join(str(iv.get(f, '')) for f in
                        ('results', 'hypotheses', 'action_iterate', 'experiments'))
        from test_intelligence import VP_DEFINITIONS
        for vp_id in self.VP_IDS:
            kws = VP_DEFINITIONS.get(vp_id, {}).get('keywords', [])
            hits = sum(1 for k in kws if k.lower() in text.lower())
            vec.append(min(1.0, hits / max(len(kws), 1)))
        # Equipment mention count (1)
        eq_count = len(iv.get('substrate_impacts', []))
        vec.append(min(1.0, eq_count / 10.0))
        # Dollar capture (1)
        total = sum(
            ei.get('annual_damage', ei.get('cost_per_event', 0) * ei.get('frequency', 1))
            for ei in iv.get('substrate_impacts', [])
            if isinstance(ei, dict)
        )
        vec.append(min(1.0, total / 500000.0))  # Normalize to $500K
        # Completeness proxy (1)
        completeness = sum(1 for f in ('results', 'hypotheses', 'action_iterate', 'experiments')
                          if iv.get(f)) / 4.0
        vec.append(completeness)
        # Text length proxy for depth (1)
        text_len = len(text)
        vec.append(min(1.0, text_len / 2000.0))
        # Total: 3+3+6+1+1+1+1 = 16 dims
        return vec


# ══════════════════════════════════════════════════════════════
# PROGRESSIVE TEST ENGINE — The Main Connector
# ══════════════════════════════════════════════════════════════
class ProgressiveTest RunEngine:
    """
    The missing connector. Ties everything into one progressive flow.

    Usage:
        engine = ProgressiveTest RunEngine(database, project_id)

        # FORWARD: Generate progressive template for next test_run
        template = engine.generate_progressive_template(person_name, test_num)

        # BACKWARD: Process completed test_run
        outputs = engine.process_completed_test_run(parsed_data)

        # REVERSE: Foreman ocular review
        review = engine.foreman_ocular(person_name)
    """

    def __init__(self, database, project_id: str):
        self.database = database
        self.project_id = project_id
        self.intel = TestIntelligence(database) if INTEL_AVAILABLE else None
        self.knn = ProgressiveKNN()

        # Build KNN from prior test_runs
        if self.intel and self.intel.test_runs:
            ioc_chain = Test RunOfChange.load_chain(project_id)
            self.knn.fit_from_chain(ioc_chain, self.intel.tests)

    # ════════════════════════════════════════════════════════
    # FORWARD: Generate progressive template
    # ════════════════════════════════════════════════════════
    def generate_progressive_template(self, person_name: str = None,
                                       test_num: int = None,
                                       q_layer: str = None,
                                       q_object: str = None) -> dict:
        """
        Generate a progressive test_run template with compound questions.
        Questions are calibrated to ALL prior tests, not just defaults.

        Returns dict with:
          - mission: focused mission statement
          - compound_tests: progressive questions (calibrated)
          - equipment_checklist: ranked equipment to probe
          - updated_hypotheses: hypothesis status + what to validate
          - knn_similar: most similar prior test_runs
          - excel_hypothesis: hypothesis from Excel plan
          - vp_summary: current VP validation status
          - reverse_gaps: what prior tests MISSED that this one should get
        """
        if not self.intel or not self.intel.state.get('ready'):
            return {'available': False, 'reason': 'No prior test_runs'}

        # Build target dict for intelligence engine
        target = {
            'name': person_name or '',
            'q_layer': q_layer or 'L2',
            'q_object': q_object or 'OC',
        }

        # Pull Excel hypothesis
        excel_hyp = {}
        if POST_GEN_AVAILABLE:
            excel_hyp = load_hypotheses_from_excel(
                self.project_id,
                person_name=person_name,
                test_num=test_num
            )
            if excel_hyp:
                target['hypothesis'] = excel_hyp.get('hypothesis_context', '')
                target['title'] = excel_hyp.get('excel_title', '')
                target['source_version'] = excel_hyp.get('excel_company', '')
                target['type'] = excel_hyp.get('excel_type', '')

        # Generate compound brief from test_intelligence
        brief = self.intel.generate_brief(target)

        # KNN: find most similar prior test_runs
        knn_similar = self.knn.find_similar(target) if SKLEARN_AVAILABLE else []

        # REVERSE GAP ANALYSIS: what did similar tests miss?
        reverse_gaps = self._analyze_reverse_gaps(knn_similar)

        # Merge everything into progressive template
        template = {
            'available': True,
            'test_num': test_num,
            'script_name': person_name,
            'project_id': self.project_id,
            'generated_at': datetime.now().isoformat(),

            # From test_intelligence
            'mission': brief.get('mission', ''),
            'compound_tests': brief.get('compound_tests', []),
            'equipment_checklist': brief.get('equipment_checklist', []),
            'updated_hypotheses': brief.get('updated_hypotheses', []),
            'vp_summary': brief.get('vp_summary', []),
            'contradictions': brief.get('contradictions', []),
            'cube_gaps': brief.get('cube_gaps', []),
            'total_damage': brief.get('total_damage', 0),
            'test_count': brief.get('test_count', 0),

            # From Excel plan
            'excel_hypothesis': excel_hyp,

            # From KNN
            'knn_similar': [
                {'script_name': iv.get('script_name', ''), 'source_version': iv.get('source_version', ''),
                 'similarity': round(sim, 3)}
                for iv, sim in knn_similar
            ],

            # Reverse gap analysis
            'reverse_gaps': reverse_gaps,

            # Current state snapshot (for IoC before)
            '_state_snapshot': copy.deepcopy(self.intel.state) if self.intel else {},
        }

        return template

    def _analyze_reverse_gaps(self, knn_similar: list) -> list:
        """
        Analyze what similar prior tests MISSED.
        Foreman ocular: find questions that should have been asked.
        """
        gaps = []

        for iv, similarity in knn_similar:
            results = iv.get('results', '')
            action = iv.get('action_iterate', '')
            text = f"{results} {action}".lower()

            # Check if they got dollar data
            has_dollars = any(c in text for c in ['$', 'cost', 'annual', 'damage'])
            if not has_dollars:
                gaps.append({
                    'type': 'DOLLAR_GAP',
                    'from_person': iv.get('script_name', ''),
                    'gap': f"{iv.get('script_name', '?')} "
                           f"at similar position had NO dollar data. "
                           f"Make sure to get specific cost numbers this time.",
                    'probe': "What is the specific annual cost of contamination damage "
                             "at your facility? Include equipment repairs, downtime, "
                             "chemical waste, and labor."
                })

            # Check if they got equipment specifics
            has_equipment = len(iv.get('substrate_impacts', [])) > 0
            if not has_equipment:
                gaps.append({
                    'type': 'EQUIPMENT_GAP',
                    'from_person': iv.get('script_name', ''),
                    'gap': f"{iv.get('script_name', '?')} had no equipment "
                           f"impact data. Probe for specific equipment this time.",
                })

            # Check if they got budget path
            has_budget = any(kw in text for kw in
                          ['budget', 'amo', 'authorize', 'would buy', 'invest'])
            if not has_budget:
                gaps.append({
                    'type': 'BUDGET_GAP',
                    'from_person': iv.get('script_name', ''),
                    'gap': f"No budget path data from similar test_run. "
                           f"Ask about purchasing authority.",
                })

        return gaps

    # ════════════════════════════════════════════════════════
    # BACKWARD: Process completed test_run
    # ════════════════════════════════════════════════════════
    def process_completed_test_run(self, parsed_data: dict,
                                     state_before: dict = None) -> dict:
        """
        Full post-test_run processing chain:
          1. Calculate intelligence AFTER (with new test)
          2. Compute before/after delta
          3. Generate BCM slide + Compound Report
          4. Save IoC snapshot
          5. Update KNN model
          6. Generate progressive questions for remaining test_runs

        Returns dict with all outputs and paths.
        """
        num = parsed_data.get('test_num', 0)
        person = parsed_data.get('script_name', 'Unknown')

        outputs = {
            'test_num': num,
            'script_name': person,
            'generated_at': datetime.now().isoformat(),
        }

        # 1. State BEFORE (from last IoC or passed in)
        if state_before is None:
            chain = Test RunOfChange.load_chain(self.project_id)
            if chain:
                state_before = chain[-1].get('state_after', {})
            elif self.intel:
                state_before = copy.deepcopy(self.intel.state)

        # 2. Recalculate intelligence WITH new test
        if self.intel:
            # Add new test to the engine and recalculate
            iv_dict = parsed_data
            if iv_dict not in self.intel.test_runs:
                self.intel.tests.append(iv_dict)
            self.intel.calculate()
            state_after = copy.deepcopy(self.intel.state)
        else:
            state_after = {}

        # 3. Compute delta
        delta = compute_delta(state_before or {}, state_after)
        outputs['delta'] = delta

        # 4. Generate slide + report (with REAL before/after)
        if POST_GEN_AVAILABLE:
            gen = PostTest RunGenerator(self.database, self.project_id)
            slide_outputs = gen.generate_all(
                parsed_data,
                intel_before=state_before,
                intel_after=state_after
            )
            outputs['slides'] = {k: str(v) for k, v in slide_outputs.items()}

        # 5. Save IoC snapshot
        ioc_path = Test RunOfChange.save(
            self.project_id, person,
            state_before or {}, state_after, delta,
            test_num=num
        )
        outputs['ioc_path'] = str(ioc_path)

        # 6. Update KNN
        if SKLEARN_AVAILABLE and self.intel:
            self.knn.fit_from_chain(
                Test RunOfChange.load_chain(self.project_id),
                self.intel.test_runs
            )

        # 7. Summary of what changed
        outputs['summary'] = self._format_delta_summary(delta, num, person)

        return outputs

    def _format_delta_summary(self, delta: dict, num: int, person: str) -> str:
        """Human-readable delta summary."""
        lines = [f"Test Run #{num} ({person}) — What Changed:"]

        # VP changes
        for vp_id, change in delta.get('vp_changes', {}).items():
            d = change['confidence_delta']
            arrow = '↑' if d > 0 else '↓' if d < 0 else '—'
            lines.append(
                f"  VP {vp_id}: {change['confidence_before']}% → "
                f"{change['confidence_after']}% ({arrow}{abs(d)}%)"
            )
            if change.get('status_changed'):
                lines.append(
                    f"    Status: {change['status_before']} → {change['status_after']}")

        # Hypothesis changes
        for h_id, change in delta.get('hypothesis_changes', {}).items():
            d = change['delta']
            lines.append(f"  Hyp {h_id}: {change['confidence_before']}% → "
                         f"{change['confidence_after']}% (Δ{d:+d}%)")

        # Damage
        dd = delta.get('damage_delta', 0)
        if dd > 0:
            lines.append(f"  Damage documented: +${dd/1000:.0f}K "
                         f"(total now ${delta.get('total_damage_after', 0)/1000:.0f}K)")

        # New equipment
        new_eq = delta.get('new_equipment', [])
        if new_eq:
            lines.append(f"  New equipment: {', '.join(new_eq)}")

        # Gaps closed
        closed = delta.get('gaps_closed', [])
        if closed:
            lines.append(f"  Cube gaps CLOSED: {', '.join(closed)}")

        # New contradictions
        new_c = delta.get('new_contradictions', [])
        if new_c:
            lines.append(f"  New contradictions: {len(new_c)}")

        return '\n'.join(lines)

    # ════════════════════════════════════════════════════════
    # REVERSE OCULAR: Foreman review
    # ════════════════════════════════════════════════════════
    def foreman_ocular(self, person_name: str = None,
                          test_num: int = None) -> dict:
        """
        Reverse ocular analysis for a completed test_run.
        What an BCM foreman would look for:
          - Were the right questions asked?
          - What was missed?
          - What should the next test_run address?
          - Is this test_run complete enough?

        Returns dict with review findings.
        """
        # Find the test_run by person name (primary) or number (fallback)
        target_iv = None
        if self.intel:
            if person_name:
                pn = person_name.strip().lower()
                for iv in self.intel.test_runs:
                    if iv.get('script_name', '').strip().lower() == pn:
                        target_iv = iv
                        break
            if not target_iv and test_num is not None:
                for iv in self.intel.test_runs:
                    if iv.get('test_num') == test_num:
                        target_iv = iv
                        break

        if not target_iv:
            return {'error': f'{person_name or test_num} not found'}

        # Get IoC state at this test_run
        state_at = Test RunOfChange.get_state_at(
            self.project_id, person_name=person_name,
            test_num=test_num)

        review = {
            'script_name': target_iv.get('script_name', ''),
            'source_version': target_iv.get('source_version', ''),
            'findings': [],
            'missed_probes': [],
            'completeness': {},
            'recommendation': '',
        }

        text = ' '.join(str(target_iv.get(f, ''))
                        for f in ('results', 'hypotheses', 'action_iterate', 'experiments'))

        # Completeness check
        fields = {
            'results': bool(target_iv.get('results', '').strip()),
            'hypotheses': bool(target_iv.get('hypotheses', '').strip()),
            'action_iterate': bool(target_iv.get('action_iterate', '').strip()),
            'experiments': bool(target_iv.get('experiments', '').strip()),
            'equipment': len(target_iv.get('substrate_impacts', [])) > 0,
            'dollar_data': any(c in text.lower() for c in ['$', 'cost', 'annual']),
            'budget_path': any(kw in text.lower() for kw in
                             ['budget', 'amo', 'authorize', 'would buy']),
        }
        review['completeness'] = fields
        score = sum(1 for v in fields.values() if v)
        review['completeness_score'] = f"{score}/{len(fields)}"

        # What was missed
        if not fields['dollar_data']:
            review['missed_probes'].append(
                "NO DOLLAR DATA: You need specific annual costs. "
                "Ask: 'What is your annual contamination damage cost?'")
        if not fields['equipment']:
            review['missed_probes'].append(
                "NO EQUIPMENT DATA: Probe specific equipment failures. "
                "Ask about chipper, screens, debarker, blow line.")
        if not fields['budget_path']:
            review['missed_probes'].append(
                "NO BUDGET PATH: You need to know if they can buy. "
                "Ask: 'What's your maintenance technology budget authority?'")

        # VP coverage check — which VPs got NO evidence from this test_run?
        from test_intelligence import VP_DEFINITIONS
        for vp_id, vp_def in VP_DEFINITIONS.items():
            hits = sum(1 for kw in vp_def['keywords'] if kw.lower() in text.lower())
            if hits == 0:
                review['findings'].append({
                    'type': 'VP_UNCOVERED',
                    'vp': vp_id,
                    'claim': vp_def['claim'],
                    'suggestion': f"No evidence for '{vp_def['claim']}'. "
                                  f"Follow up with targeted question.",
                })

        # Foreman recommendation
        if score < 4:
            review['recommendation'] = (
                f"INCOMPLETE ({review['completeness_score']}). "
                f"Schedule follow-up to fill: "
                f"{', '.join(k for k, v in fields.items() if not v)}")
        elif len(review['missed_probes']) > 1:
            review['recommendation'] = (
                f"ADEQUATE but gaps. {len(review['missed_probes'])} probes missed. "
                f"Next test_run should target these gaps.")
        else:
            review['recommendation'] = "SOLID test_run. All major areas covered."

        return review

    # ════════════════════════════════════════════════════════
    # EXCEL PLAN UPDATE — Progressive questions
    # ════════════════════════════════════════════════════════
    def update_excel_plan_progressive(self) -> List[dict]:
        """
        Generate progressive updates for the Excel test_run plan.
        For each pending test_run, generates compound questions
        calibrated to ALL prior tests.

        Returns list of dicts with updated questions per test.
        """
        if not self.intel or not self.intel.state.get('ready'):
            return []

        completed_names = set(
            iv.get('script_name', '').strip().lower() for iv in self.intel.test_runs
            if iv.get('results') and '[PENDING' not in str(iv.get('results', ''))
        )

        updates = []
        # Get all test_run plans from Excel
        excel_path = find_project_config(self.project_id) if PATHS_AVAILABLE else None
        if not excel_path:
            return []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
            ws = wb.active

            for r in range(2, ws.max_row + 1):
                row_num = ws.cell(r, 1).value
                if row_num is None:
                    continue

                row_name = str(ws.cell(r, 2).value or "")
                if row_name.strip().lower() in completed_names:
                    continue
                row_title = str(ws.cell(r, 3).value or "")
                row_type = str(ws.cell(r, 4).value or "")
                row_company = str(ws.cell(r, 5).value or "")
                row_hyp = str(ws.cell(r, 6).value or "")

                # Parse Q-Cube position from hypothesis text or default
                q_layer = 'L2'
                q_object = 'OC'
                for layer in ['L1', 'L2', 'L3']:
                    if layer in row_type.upper() or layer in row_hyp:
                        q_layer = layer
                for obj in ['OA', 'OB', 'OC']:
                    if obj in row_type.upper() or obj in row_hyp:
                        q_object = obj

                # Generate progressive template for this person
                template = self.generate_progressive_template(
                    test_num=row_num,
                    person_name=row_name,
                    q_layer=q_layer,
                    q_object=q_object,
                )

                updates.append({
                    'row': r,
                    'test_num': row_num,
                    'script_name': row_name,
                    'source_version': row_company,
                    'progressive_questions': template.get('compound_tests', []),
                    'updated_hypotheses': template.get('updated_hypotheses', []),
                    'knn_similar': template.get('knn_similar', []),
                    'reverse_gaps': template.get('reverse_gaps', []),
                    'vp_status': template.get('vp_summary', []),
                })

            wb.close()
        except Exception as e:
            print(f"  [WARN] Excel plan update failed: {e}")

        return updates


# ══════════════════════════════════════════════════════════════
# STANDALONE TEST
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 60)
    print("PROGRESSIVE TEST ENGINE — Test")
    print("=" * 60)

    # Load from Baseline deck
    deck_path = Path("EPIC_Baseline_test_database_deck.json")
    for p in [deck_path, Path(__file__).parent / deck_path.name]:
        if p.exists():
            deck_path = p
            break

    if deck_path.exists():
        with open(deck_path, 'r') as f:
            deck = json.load(f)

        engine = ProgressiveTest RunEngine.__new__(ProgressiveTest RunEngine)
        engine.database = None
        engine.project_id = "BCM_SUBSTRATE"

        if INTEL_AVAILABLE:
            engine.intel = TestIntelligence()
            engine.intel.load_json(str(deck_path))
            engine.intel.calculate()
        else:
            engine.intel = None

        engine.knn = ProgressiveKNN()
        if SKLEARN_AVAILABLE and engine.intel:
            engine.knn.fit_from_chain([], engine.intel.tests)

        print(f"\nLoaded {len(engine.intel.tests)} test_runs")

        # Test FORWARD: generate progressive template for test_run #14
        print("\n" + "=" * 60)
        print("FORWARD: Progressive template for test_run #14 (SAPPI)")
        print("=" * 60)
        template = engine.generate_progressive_template(
            person_name="Marty Richards", test_num=14,
            q_layer="L2", q_object="OC"
        )
        print(f"  Mission: {template['mission'][:100]}...")
        print(f"  Compound Questions ({len(template['compound_tests'])}):")
        for i, q in enumerate(template['compound_tests'], 1):
            print(f"    Q{i}. {q[:100]}...")
        print(f"  KNN Similar: {template['knn_similar']}")
        print(f"  Reverse Gaps: {len(template['reverse_gaps'])}")

        # Test REVERSE OCULAR: review test_run #2
        print("\n" + "=" * 60)
        print("REVERSE: Foreman ocular for test_run #2 (Ryan Thorton)")
        print("=" * 60)
        review = engine.foreman_ocular(person_name="Ryan Thorton")
        print(f"  Completeness: {review.get('completeness_score')}")
        print(f"  Missed probes: {len(review.get('missed_probes', []))}")
        for mp in review.get('missed_probes', []):
            print(f"    - {mp[:80]}...")
        print(f"  Recommendation: {review.get('recommendation')}")
    else:
        print(f"Deck not found: {deck_path}")
