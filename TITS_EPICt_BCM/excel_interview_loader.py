# -*- coding: utf-8 -*-
#!/usr/bin/env python3
"""
EXCEL TEST LOADER - Living Document Integration
====================================================
Loads test_run plans from BCM_Test_Plan.xlsx

THE EXCEL FILE IS THE SINGLE SOURCE OF TRUTH.
Grad students edit Excel â†’ App reflects automatically.
No more hardcoding tests in Python.

Expected Excel Structure (BCM_Test_Plan.xlsx):
- Sheet: "FUSION Test Plans"
- Rows 0-8: Header section (Value Proposition)
- Row 9+: Test Run data
- Columns: #, Name, Title, Type, Company, Hypothesis, Questions

Author: GIBUSH AI Engineering Team
"""

import sys
from pathlib import Path
from typing import Dict, List, Optional

# Try to import pandas
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("Warning: pandas not installed. Excel loading disabled.")
    print("Install with: pip install pandas openpyxl")


# ============================================================================
# CONFIGURATION
# ============================================================================

# Excel file names by project (the living documents)
EXCEL_FILENAMES = {
    'BCM_SUBSTRATE': 'BCM_Test_Plan.xlsx',
    'BCM_NAVIGATION': 'AISOS_SPINE_BCM_Test_Plan.xlsx',
    'default': 'BCM_Test_Plan.xlsx'
}

# Current active project (can be changed at runtime)
_active_project = 'BCM_SUBSTRATE'

def set_active_project(project_id: str):
    """Set which project's Excel file to load"""
    global _active_project
    if project_id in EXCEL_FILENAMES:
        _active_project = project_id
    else:
        _active_project = 'default'

def get_active_project() -> str:
    """Get current active project"""
    return _active_project

def get_excel_filename() -> str:
    """Get Excel filename for current project"""
    return EXCEL_FILENAMES.get(_active_project, EXCEL_FILENAMES['default'])

# Search paths for the Excel file
def get_search_paths() -> List[Path]:
    """Get list of paths to search for Excel file - project_team_name_test_run_xcel ONLY"""
    paths = []
    
    # project_team_name_test_run_xcel - THE SINGLE SOURCE
    paths.append(Path.cwd() / "project_team_name_test_run_xcel")
    
    # Windows TITS path
    if sys.platform == "win32":
        paths.append(Path("C:/TITS/TITS_Production/gibush_icorps/project_team_name_test_run_xcel"))
    
    return paths


def find_excel_file(project_id: str = None) -> Optional[Path]:
    """
    Find the Excel file in known locations.
    
    Args:
        project_id: Optional project ID. If None, uses active project.
    
    Returns:
        Path to Excel file if found, None otherwise
    """
    if project_id:
        filename = EXCEL_FILENAMES.get(project_id, EXCEL_FILENAMES['default'])
    else:
        filename = get_excel_filename()
    
    for search_path in get_search_paths():
        excel_path = search_path / filename
        if excel_path.exists():
            return excel_path
    
    return None


def get_project_folder() -> Path:
    """Get the project_team_name_test_run_xcel folder path"""
    if sys.platform == "win32":
        return Path("C:/TITS/TITS_Production/gibush_icorps/project_team_name_test_run_xcel")
    return Path.cwd() / "project_team_name_test_run_xcel"


def list_projects() -> List[str]:
    """List all registered project IDs"""
    return [k for k in EXCEL_FILENAMES.keys() if k != 'default']


def create_project(project_id: str, excel_filename: str = None) -> Optional[Path]:
    """
    Create a new project with blank Excel template.
    
    Args:
        project_id: Unique project identifier (e.g., 'NEW_PROJECT')
        excel_filename: Excel filename. If None, uses {project_id}_BCM_Test_Plan.xlsx
    
    Returns:
        Path to created Excel file, or None if failed
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("Error: openpyxl required. pip install openpyxl")
        return None
    
    # Generate filename if not provided
    if excel_filename is None:
        excel_filename = f"{project_id}_BCM_Test_Plan.xlsx"
    
    # Get folder path
    folder = get_project_folder()
    folder.mkdir(parents=True, exist_ok=True)
    
    excel_path = folder / excel_filename
    
    # Don't overwrite existing
    if excel_path.exists():
        print(f"Project Excel already exists: {excel_path}")
        # Just register it
        EXCEL_FILENAMES[project_id] = excel_filename
        return excel_path
    
    # Create workbook with template structure
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FUSION Test Plans"
    
    # Header style
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF")
    
    # Value Proposition section (rows 1-8)
    vp_labels = [
        ("A1", "PROJECT:", "B1", project_id),
        ("A2", "THE GAP WE FILL:", "B2", ""),
        ("A3", "WHAT THEY CAN'T DO:", "B3", ""),
        ("A4", "SYNERGY PLAY:", "B4", ""),
        ("A5", "DIFFERENTIATION:", "B5", ""),
        ("A6", "", "B6", ""),
        ("A7", "", "B7", ""),
        ("A8", "", "B8", ""),
    ]
    
    for label_cell, label, value_cell, value in vp_labels:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = value
    
    # Column headers (row 9)
    headers = ["#", "Name", "Title", "Type", "Company", "Hypothesis", "Questions"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 60
    
    # Add one blank test_run row as template
    ws.cell(row=10, column=1, value=1)
    ws.cell(row=10, column=2, value="[Name]")
    ws.cell(row=10, column=3, value="[Title]")
    ws.cell(row=10, column=4, value="[Type]")
    ws.cell(row=10, column=5, value="[Company]")
    ws.cell(row=10, column=6, value="[Hypothesis]")
    ws.cell(row=10, column=7, value="[Questions]")
    
    # Save
    wb.save(excel_path)
    
    # Register project
    EXCEL_FILENAMES[project_id] = excel_filename
    
    print(f"Created project: {project_id}")
    print(f"Excel: {excel_path}")
    
    return excel_path


# ============================================================================
# EXCEL LOADER CLASS
# ============================================================================

class JsonTestLoader:
    """
    Loads test_run plan data from Excel spreadsheet.
    
    The Excel file is the SINGLE SOURCE OF TRUTH.
    Supports multiple projects (BCM_SUBSTRATE, BCM_NAVIGATION).
    """
    
    # Excel structure constants
    HEADER_ROWS = 9  # Rows 0-8 are header/VP section
    SHEET_NAME = "FUSION Test Plans"
    
    # Column mapping (based on Excel structure)
    COL_NUM = 0       # Test Run number
    COL_NAME = 1      # Person name
    COL_TITLE = 2     # Job title
    COL_TYPE = 3      # Customer type
    COL_COMPANY = 4   # Company name (or Department for internal)
    COL_HYPOTHESIS = 5  # Hypothesis
    COL_QUESTIONS = 6   # Questions
    
    def __init__(self, excel_path: Path = None, project_id: str = None):
        """
        Initialize loader.
        
        Args:
            excel_path: Path to Excel file. If None, searches known locations.
            project_id: Project ID (BCM_SUBSTRATE or BCM_NAVIGATION). If None, uses active.
        """
        self.project_id = project_id or get_active_project()
        self.excel_path = excel_path or find_excel_file(self.project_id)
        self._cache = None
        self._vp_cache = None
    
    @property
    def is_available(self) -> bool:
        """Check if Excel file is available"""
        return self.excel_path is not None and self.excel_path.exists()
    
    def load_test_runs(self, force_reload: bool = False) -> List[Dict]:
        """
        Load test_run plans from Excel.
        
        Args:
            force_reload: If True, bypass cache and reload from disk
            
        Returns:
            List of test_run dicts with keys: num, name, title, type, company, hypothesis, questions
        """
        if not PANDAS_AVAILABLE:
            print("Error: pandas not available")
            return []
        
        if not self.is_available:
            print(f"Error: Excel file not found. Searched: {[str(p) for p in get_search_paths()]}")
            return []
        
        if self._cache is not None and not force_reload:
            return self._cache
        
        try:
            # Read Excel with no header (we'll parse manually)
            df = pd.read_excel(self.excel_path, sheet_name=0, header=None)
            
            tests = []
            
            # Start after header section
            for idx in range(self.HEADER_ROWS, len(df)):
                row = df.iloc[idx]
                
                # Skip if no name
                if pd.isna(row[self.COL_NAME]) or str(row[self.COL_NAME]).strip() == '':
                    continue

                # Parse test_run number
                # Formula cells (=A10+1) evaluate to NaN on Windows — derive from row position
                raw_num = row[self.COL_NUM]
                if pd.isna(raw_num) or str(raw_num).strip().startswith('='):
                    num = idx - self.HEADER_ROWS + 1
                else:
                    try:
                        num = int(float(str(raw_num).strip()))
                    except (ValueError, TypeError):
                        num = idx - self.HEADER_ROWS + 1
                
                # Clean up questions - replace \\n with actual newlines
                questions = str(row[self.COL_QUESTIONS]) if pd.notna(row[self.COL_QUESTIONS]) else ''
                questions = questions.replace('\\n', '\n')
                
                test_run = {
                    'num': num,
                    'name': str(row[self.COL_NAME]) if pd.notna(row[self.COL_NAME]) else '',
                    'title': str(row[self.COL_TITLE]) if pd.notna(row[self.COL_TITLE]) else '',
                    'type': str(row[self.COL_TYPE]) if pd.notna(row[self.COL_TYPE]) else '',
                    'source_version': str(row[self.COL_COMPANY]) if pd.notna(row[self.COL_COMPANY]) else '',
                    'hypothesis': str(row[self.COL_HYPOTHESIS]) if pd.notna(row[self.COL_HYPOTHESIS]) else '',
                    'questions': questions
                }
                
                tests.append(test_run)
            
            # Sort by test_run number
            tests.sort(key=lambda x: x['num'])
            
            self._cache = test_runs
            return test_runs
            
        except Exception as e:
            print(f"Error loading Excel: {e}")
            return []
    
    def load_value_proposition(self, force_reload: bool = False) -> Dict:
        """
        Load value proposition from Excel header section.
        
        Returns dict with keys: gap_we_fill, what_they_cant_do, synergy_play, differentiation
        """
        if not PANDAS_AVAILABLE:
            return {}
        
        if not self.is_available:
            return {}
        
        if self._vp_cache is not None and not force_reload:
            return self._vp_cache
        
        try:
            df = pd.read_excel(self.excel_path, sheet_name=0, header=None)
            
            vp = {
                'gap_we_fill': '',
                'what_they_cant_do': '',
                'synergy_play': '',
                'differentiation': ''
            }
            
            # Parse header rows (0-8)
            for idx in range(min(9, len(df))):
                row = df.iloc[idx]
                label = str(row[0]).upper() if pd.notna(row[0]) else ''
                value = str(row[2]) if pd.notna(row[2]) else ''
                
                if 'GAP WE FILL' in label:
                    vp['gap_we_fill'] = value
                elif "CAN'T DO" in label or "CANT DO" in label:
                    vp['what_they_cant_do'] = value
                elif 'SYNERGY' in label:
                    vp['synergy_play'] = value
                elif 'DIFFERENTIATION' in label:
                    vp['differentiation'] = value
            
            self._vp_cache = vp
            return vp
            
        except Exception as e:
            print(f"Error loading VP from Excel: {e}")
            return {}
    
    def reload(self):
        """Force reload from disk (clears cache)"""
        self._cache = None
        self._vp_cache = None
        return self.load_test_runs(force_reload=True)
    
    def get_test_run_by_num(self, num: int) -> Optional[Dict]:
        """Get specific test_run by number"""
        tests = self.load_test_runs()
        for test_entry in test_runs:
            if test_entry['num'] == num:
                return test_run
        return None


# ============================================================================
# MODULE-LEVEL CONVENIENCE FUNCTIONS
# ============================================================================

# Global loader instances by project
_loaders = {}

def get_loader(project_id: str = None) -> JsonTestLoader:
    """Get or create loader instance for specified project"""
    global _loaders
    pid = project_id or get_active_project()
    if pid not in _loaders:
        _loaders[pid] = JsonTestLoader(project_id=pid)
    return _loaders[pid]


def load_test_plans(project_id: str = None) -> List[Dict]:
    """
    Load test_run plans from Excel.
    
    This is the main entry point for the validation_test_collector.py
    """
    return get_loader(project_id).load_test_runs()


def load_value_proposition(project_id: str = None) -> Dict:
    """Load value proposition from Excel header"""
    return get_loader(project_id).load_value_proposition()


def reload_from_excel(project_id: str = None) -> List[Dict]:
    """Force reload from Excel (bypass cache AND recreate loader)"""
    global _loaders
    pid = project_id or get_active_project()
    
    # FORCE new loader creation - don't reuse cached loader with wrong path
    _loaders[pid] = JsonTestLoader(project_id=pid)
    
    return _loaders[pid].load_test_runs(force_reload=True)


def get_excel_path(project_id: str = None) -> Optional[Path]:
    """Get path to Excel file (for display in UI)"""
    loader = get_loader(project_id)
    return loader.excel_path if loader.is_available else None


def is_excel_available(project_id: str = None) -> bool:
    """Check if Excel file is available"""
    return get_loader(project_id).is_available


# ============================================================================
# MAIN (for testing)
# ============================================================================

if __name__ == "__main__":
    print("Excel Test Run Loader Test")
    print("=" * 60)
    
    # Test BCM_SUBSTRATE (GIBUSH)
    print("\nðŸ“ BCM_SUBSTRATE (GIBUSH - External BCM)")
    set_active_project('BCM_SUBSTRATE')
    loader1 = get_loader('BCM_SUBSTRATE')
    
    if loader1.is_available:
        print(f"   Excel found: {loader1.excel_path.name}")
        tests = loader1.load_test_runs()
        print(f"   Loaded {len(tests)} test_runs")
        for iv in test_runs[:3]:
            print(f"     #{iv['num']}: {iv['name']} - {iv['source_version']}")
    else:
        print("   Excel file NOT found")
    
    # Test BCM_NAVIGATION (AISOS)
    print("\nðŸ›¡ï¸  BCM_NAVIGATION (AISOS - Internal Safety)")
    set_active_project('BCM_NAVIGATION')
    loader2 = get_loader('BCM_NAVIGATION')
    
    if loader2.is_available:
        print(f"   Excel found: {loader2.excel_path.name}")
        tests = loader2.load_test_runs()
        print(f"   Loaded {len(tests)} test_runs")
        
        # Show by position type
        types = {}
        for iv in test_runs:
            t = iv.get('type', 'Unknown')
            types[t] = types.get(t, 0) + 1
        print("   By position type:")
        for t, count in sorted(types.items()):
            print(f"     {t}: {count}")
    else:
        print("   Excel file NOT found")
    
    print("\n" + "=" * 60)
